from dash import dcc, html, dash_table, Dash
from dash.dependencies import ClientsideFunction, Input, Output, State
from datetime import datetime, timedelta
from db_connection import get_db_connection
from flask_caching import Cache
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

cache = Cache(config={'CACHE_TYPE': 'simple'})
class AnalysisRegister:
    def __init__(self, flask_app):
        self.app = Dash(__name__, server=flask_app,
                        url_base_pathname='/analysis_register/')
        self.setup_layout()
        self.setup_callbacks()

    def get_data(self):
        conn = get_db_connection()
        cursor = conn.cursor()
        sql = """SELECT AnlysID,SampleID,LabID,TestType,UserID,Material,M_C,O_C,FFA,FM,SS,PROTEIN,CLR,MIV,EO,IV,SV,Date_Time_Stmp FROM analysisreg
                ORDER BY Date_Time_Stmp DESC;
                """
        cursor.execute(sql)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        cursor.close()
        conn.close()
        data = [dict(zip(columns, row)) for row in rows]
        return data
  
    def setup_layout(self):
        data = self.get_data()

        columns = [{"name": col, "id": col} for col in data[0].keys()]

        self.app.layout = html.Div([
            html.Link(
                rel='stylesheet',
                href='https://cdnjs.cloudflare.com/ajax/libs/bootstrap-icons/1.4.0/font/bootstrap-icons.min.css'
            ),
            html.Div([
                html.Div(className='button-container', children=[
                    html.A([html.I(className="bi bi-house"), 'Home'],
                           href='/home', className='link-button'),
                    html.Div(id='dummy-output'),
                    html.Button([html.I(className="bi bi-file-excel"), " Download Excel"],
                                id="p-button", className='p-button'),
                    dcc.Download(id="download-dataframe-xlsx"),
                ]),
                html.H1('Analysis Register',
                        style={'textAlign': 'center', 'fontWeight': 'bold', 'flex': '1'}),
                html.Div([html.Label('Date:', style={
                    'margin-right': '10px', 'fontWeight': 'bold'}),
                    html.Label(datetime.now().strftime('%Y-%m-%d'),
                               style={'margin-right': '20px'}),], style={'display': 'inline', 'fontWeight': 'bold', 'border-style': 'solid',
                                                                         'border-color': 'black',
                                                                         'backgroundColor': 'white',
                                                                         'padding': '5px 10px',
                                                                         'borderRadius': '5px',
                                                                         'fontSize': '20px',
                                                                         'margin': '10px'
                                                                         })
            ], style={'display': 'flex', 'align-items': 'center', 'justify-content': 'space-between', 'width': '100%', 'margin': '10px'}),
            html.Div(id='box', children=[
                html.Label('Select Report Date:', style={
                    'margin-right': '10px', 'fontWeight': 'bold'}),
                dcc.DatePickerRange(
                    id='date-picker-Range',
                    min_date_allowed=datetime(2020, 3, 28),
                    max_date_allowed=datetime.now(),
                    start_date=datetime.now().date() - timedelta(days=1),
                    end_date=datetime.now().date(),
                    display_format='YYYY-MM-DD',
                    style={
                        'display': 'inline-block',
                        'fontWeight': 'bold',
                        'backgroundColor': 'white',
                    },
                ),
            ], style={'display': 'inline-flex', 'justifyContent': 'left', 'alignItems': 'center', 'margin': '10px', 'border-style': 'solid',
                      'border-color': 'black', 'padding': '5px 10px',
                      'borderRadius': '5px', 'fontSize': '20px',
                      'margin': '10px'}),
            dash_table.DataTable(
                id='table',
                columns=columns,
                data=data[:1000],
                merge_duplicate_headers=True,
                page_size=1000,
                style_cell={
                    'textAlign': 'center',
                    'minWidth': '10px',
                    'width': '50px',
                    'maxWidth': '150px',
                    'whiteSpace': 'normal',
                },
                style_header={
                    'backgroundColor': 'rgb(230, 230, 230)',
                    'fontWeight': 'bold'
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': 'rgb(248, 248, 248)'
                    },
                ]
            )
        ])

    def setup_callbacks(self):
        @self.app.callback(
            Output('table', 'data'),
            [
                Input('date-picker-Range', 'start_date'),
                Input('date-picker-Range', 'end_date')
            ]
        )
        def update_table(start_date, end_date):
            data = self.get_data()
            if start_date is not None and end_date is not None:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                filtered_data = []
                for row in data:
                    row_date = row['Date_Time_Stmp']
                    if start_date <= row_date < end_date:
                        filtered_data.append(row) 
                        continue
                data = filtered_data
            return data
        @self.app.callback(
            Output("download-dataframe-xlsx", "data"),
            [
                Input('p-button', 'n_clicks'),
                State('date-picker-Range', 'start_date'),
                State('date-picker-Range', 'end_date')
            ],
            prevent_initial_call=True,
        )
        def func(n_clicks, start_date, end_date):
            data = self.get_data()
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                filtered_data = []
                for row in data:
                    row_date = row['Date_Time_Stmp']
                    if start_date <= row_date < end_date:
                        filtered_data.append(row)
                data = filtered_data

            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Analysis Register"
            worksheet.merge_cells('A1:R1')
            worksheet['A1'].value = "Analysis Register"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

            for r_idx, row in enumerate(dataframe_to_rows(pd.DataFrame(data), index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            for col in range(1, len(data[0]) + 1):
                col_letter = get_column_letter(col)
                header = list(data[0].keys())[col - 1]

                worksheet[f'{col_letter}2'].value = header
                worksheet[f'{col_letter}2'].font = Font(bold=True)
                worksheet[f'{col_letter}2'].alignment = openpyxl.styles.Alignment(horizontal='center')
            last_col_letter = get_column_letter(len(data[0]))
            worksheet.column_dimensions[last_col_letter].width = 20   

            workbook.save(output)
            output.seek(0)
            return dcc.send_bytes(output.getvalue(), "Analysis_Register.xlsx")

        
