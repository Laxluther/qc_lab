
from dash import dcc, html, dash_table, Dash
from dash.dependencies import ClientsideFunction, Input, Output, State
from datetime import datetime, timedelta
from db_connection import get_db_connection
from dash_extensions.javascript import assign
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

print_js = assign("function(n_clicks){window.print();}")


class DailyReportDashboardMonthly:
    def __init__(self, flask_app):
        self.app = Dash(__name__, server=flask_app,
                        url_base_pathname='/daily_analysis_report_monthly/')
        self.setup_layout()
        self.setup_callbacks()

    def get_data(self):
        conn = get_db_connection()
        cursor = conn.cursor()
        sql = """SELECT s.SampleID,s.Date_Time, a.TestType, a.Material, a.M_C, a.O_C, a.FFA, a.CLR, a.MIV, a.EO, a.IV, a.SV
                FROM SampleReg s
                INNER JOIN AnalysisReg a ON s.sampleid = a.sampleid WHERE s.SampleType = 'Running'
                ORDER BY s.Date_Time DESC;
                """
        cursor.execute(sql)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        cursor.close()
        conn.close()
        return [dict(zip(columns, row)) for row in rows]

    def format_dataframe_for_excel(self, reshaped_data):
        df = pd.DataFrame(reshaped_data)

        df.columns = pd.MultiIndex.from_tuples([
            ("", "", "Date/Time"),
            ("Mustard Seed", "", "Sample ID"),
            ("Mustard Seed", "NIR", "M/C"),
            ("Mustard Seed", "NIR", "O/C"),
            ("Mustard Seed", "MANUAL", "M/C"),
            ("Mustard Seed", "MANUAL", "O/C"),
            ("Mustard Cake", "", "Sample ID"),
            ("Mustard Cake", "NIR", "M/C"),
            ("Mustard Cake", "NIR", "O/C"),
            ("Mustard Cake", "MANUAL", "M/C"),
            ("Mustard Cake", "MANUAL", "O/C"),
            ("Mustard Ghani", "", "Sample ID"),
            ("Mustard Ghani", "NIR", "M/C"),
            ("Mustard Ghani", "NIR", "O/C"),
            ("Mustard Ghani", "MANUAL", "M/C"),
            ("Mustard Ghani", "MANUAL", "O/C"),
            ("KGMO", "", "Sample ID"),
            ("KGMO", "MANUAL", "FFA"),
            ("KGMO", "MANUAL", "CLR"),
            ("KGMO", "MANUAL", "MIV"),
            ("KGMO", "MANUAL", "EO"),
            ("KGMO", "MANUAL", "IV"),
            ("KGMO", "MANUAL", "SV"),
        ])
        return df

    def calculate_footer_data(self, data):
        footer = {
            'Date/Time': 'Total/Average',
            'Mustard Seed Sample ID': sum(1 for row in data if row['Mustard Seed Sample ID']),
            'Mustard Seed NIR M/C': round(sum(row['Mustard Seed NIR M/C'] for row in data if row['Mustard Seed NIR M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Seed NIR M/C'] is not None)), 2),
            'Mustard Seed NIR O/C': round(sum(row['Mustard Seed NIR O/C'] for row in data if row['Mustard Seed NIR O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Seed NIR O/C'] is not None)), 2),
            'Mustard Seed MANUAL M/C': round(sum(row['Mustard Seed MANUAL M/C'] for row in data if row['Mustard Seed MANUAL M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Seed MANUAL M/C'] is not None)), 2),
            'Mustard Seed MANUAL O/C': round(sum(row['Mustard Seed MANUAL O/C'] for row in data if row['Mustard Seed MANUAL O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Seed MANUAL O/C'] is not None)), 2),
            'Mustard Cake Sample ID': sum(1 for row in data if row['Mustard Cake Sample ID']),
            'Mustard Cake NIR M/C': round(sum(row['Mustard Cake NIR M/C'] for row in data if row['Mustard Cake NIR M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Cake NIR M/C'] is not None)), 2),
            'Mustard Cake NIR O/C': round(sum(row['Mustard Cake NIR O/C'] for row in data if row['Mustard Cake NIR O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Cake NIR O/C'] is not None)), 2),
            'Mustard Cake MANUAL M/C': round(sum(row['Mustard Cake MANUAL M/C'] for row in data if row['Mustard Cake MANUAL M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Cake MANUAL M/C'] is not None)), 2),
            'Mustard Cake MANUAL O/C': round(sum(row['Mustard Cake MANUAL O/C'] for row in data if row['Mustard Cake MANUAL O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Cake MANUAL O/C'] is not None)), 2),
            'Mustard Ghani Sample ID': sum(1 for row in data if row['Mustard Ghani Sample ID']),
            'Mustard Ghani NIR M/C': round(sum(row['Mustard Ghani NIR M/C'] for row in data if row['Mustard Ghani NIR M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Ghani NIR M/C'] is not None)), 2),
            'Mustard Ghani NIR O/C': round(sum(row['Mustard Ghani NIR O/C'] for row in data if row['Mustard Ghani NIR O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Ghani NIR O/C'] is not None)), 2),
            'Mustard Ghani MANUAL M/C': round(sum(row['Mustard Ghani MANUAL M/C'] for row in data if row['Mustard Ghani MANUAL M/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Ghani MANUAL M/C'] is not None)), 2),
            'Mustard Ghani MANUAL O/C': round(sum(row['Mustard Ghani MANUAL O/C'] for row in data if row['Mustard Ghani MANUAL O/C'] is not None) / max(1, sum(1 for row in data if row['Mustard Ghani MANUAL O/C'] is not None)), 2),
            'KGMO Sample ID': sum(1 for row in data if row['KGMO Sample ID']),
            'KGMO MANUAL FFA': round(sum(row['KGMO MANUAL FFA'] for row in data if row['KGMO MANUAL FFA'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL FFA'] is not None)), 2),
            'KGMO MANUAL CLR': round(sum(row['KGMO MANUAL CLR'] for row in data if row['KGMO MANUAL CLR'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL CLR'] is not None)), 2),
            'KGMO MANUAL MIV': round(sum(row['KGMO MANUAL MIV'] for row in data if row['KGMO MANUAL MIV'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL MIV'] is not None)), 2),
            'KGMO MANUAL EO': round(sum(row['KGMO MANUAL EO'] for row in data if row['KGMO MANUAL EO'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL EO'] is not None)), 2),
            'KGMO MANUAL IV': round(sum(row['KGMO MANUAL IV'] for row in data if row['KGMO MANUAL IV'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL IV'] is not None)), 2),
            'KGMO MANUAL SV': round(sum(row['KGMO MANUAL SV'] for row in data if row['KGMO MANUAL SV'] is not None) / max(1, sum(1 for row in data if row['KGMO MANUAL SV'] is not None)), 2),
        }
        return footer

    def append_footer_data(self, data):
        footer = self.calculate_footer_data(data)
        data.append(footer)
        return data

    def reshape_data(self, data):
        reshaped_data = {}
        for row in data:
            key = (row['SampleID'], row['Date_Time'])
            if key not in reshaped_data:
                reshaped_data[key] = {
                    'Date/Time': row['Date_Time'].strftime('%Y-%m-%d %H:%M'),
                    'Mustard Seed Sample ID': None,
                    'Mustard Seed NIR M/C': None,
                    'Mustard Seed NIR O/C': None,
                    'Mustard Seed MANUAL M/C': None,
                    'Mustard Seed MANUAL O/C': None,
                    'Mustard Cake Sample ID': None,
                    'Mustard Cake NIR M/C': None,
                    'Mustard Cake NIR O/C': None,
                    'Mustard Cake MANUAL M/C': None,
                    'Mustard Cake MANUAL O/C': None,
                    'Mustard Ghani Sample ID': None,
                    'Mustard Ghani NIR M/C': None,
                    'Mustard Ghani NIR O/C': None,
                    'Mustard Ghani MANUAL M/C': None,
                    'Mustard Ghani MANUAL O/C': None,
                    'KGMO Sample ID': None,
                    'KGMO MANUAL FFA': None,
                    'KGMO MANUAL CLR': None,
                    'KGMO MANUAL MIV': None,
                    'KGMO MANUAL EO': None,
                    'KGMO MANUAL IV': None,
                    'KGMO MANUAL SV': None,
                }
            if row['Material'] == 'MUSTARD SEED':
                reshaped_data[key]['Mustard Seed Sample ID'] = row['SampleID']
                if row['TestType'] == 'NIR':
                    reshaped_data[key]['Mustard Seed NIR M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Seed NIR O/C'] = row['O_C']
                elif row['TestType'] == 'MANUAL':
                    reshaped_data[key]['Mustard Seed MANUAL M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Seed MANUAL O/C'] = row['O_C']
            elif row['Material'] == 'MUSTARD CAKE':
                reshaped_data[key]['Mustard Cake Sample ID'] = row['SampleID']
                if row['TestType'] == 'NIR':
                    reshaped_data[key]['Mustard Cake NIR M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Cake NIR O/C'] = row['O_C']
                elif row['TestType'] == 'MANUAL':
                    reshaped_data[key]['Mustard Cake MANUAL M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Cake MANUAL O/C'] = row['O_C']
            elif row['Material'] == 'MUSTARD GHANI':
                reshaped_data[key]['Mustard Ghani Sample ID'] = row['SampleID']
                if row['TestType'] == 'NIR':
                    reshaped_data[key]['Mustard Ghani NIR M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Ghani NIR O/C'] = row['O_C']
                elif row['TestType'] == 'MANUAL':
                    reshaped_data[key]['Mustard Ghani MANUAL M/C'] = row['M_C']
                    reshaped_data[key]['Mustard Ghani MANUAL O/C'] = row['O_C']
            elif row['Material'] == 'KGMO':
                reshaped_data[key]['KGMO Sample ID'] = row['SampleID']
                reshaped_data[key]['KGMO MANUAL FFA'] = row['FFA']
                reshaped_data[key]['KGMO MANUAL CLR'] = row['CLR']
                reshaped_data[key]['KGMO MANUAL MIV'] = row['MIV']
                reshaped_data[key]['KGMO MANUAL EO'] = row['EO']
                reshaped_data[key]['KGMO MANUAL IV'] = row['IV']
                reshaped_data[key]['KGMO MANUAL SV'] = row['SV']
        return list(reshaped_data.values())

    def setup_layout(self):
        data = self.get_data()
        reshaped_data = self.reshape_data(data)
        reshaped_data = self.append_footer_data(reshaped_data)

        columns = [
            {"name": ["", "", "Date/Time"], "id": "Date/Time"},
            {"name": ["Mustard Seed", "", "Sample ID"],
                "id": "Mustard Seed Sample ID"},
            {"name": ["Mustard Seed", "NIR", "M/C"],
                "id": "Mustard Seed NIR M/C"},
            {"name": ["Mustard Seed", "NIR", "O/C"],
                "id": "Mustard Seed NIR O/C"},
            {"name": ["Mustard Seed", "MANUAL", "M/C"],
                "id": "Mustard Seed MANUAL M/C"},
            {"name": ["Mustard Seed", "MANUAL", "O/C"],
                "id": "Mustard Seed MANUAL O/C"},
            {"name": ["Mustard Cake", "", "Sample ID"],
                "id": "Mustard Cake Sample ID"},
            {"name": ["Mustard Cake", "NIR", "M/C"],
                "id": "Mustard Cake NIR M/C"},
            {"name": ["Mustard Cake", "NIR", "O/C"],
                "id": "Mustard Cake NIR O/C"},
            {"name": ["Mustard Cake", "MANUAL", "M/C"],
                "id": "Mustard Cake MANUAL M/C"},
            {"name": ["Mustard Cake", "MANUAL", "O/C"],
                "id": "Mustard Cake MANUAL O/C"},
            {"name": ["Mustard Ghani", "", "Sample ID"],
                "id": "Mustard Ghani Sample ID"},
            {"name": ["Mustard Ghani", "NIR", "M/C"],
                "id": "Mustard Ghani NIR M/C"},
            {"name": ["Mustard Ghani", "NIR", "O/C"],
                "id": "Mustard Ghani NIR O/C"},
            {"name": ["Mustard Ghani", "MANUAL", "M/C"],
                "id": "Mustard Ghani MANUAL M/C"},
            {"name": ["Mustard Ghani", "MANUAL", "O/C"],
                "id": "Mustard Ghani MANUAL O/C"},
            {"name": ["KGMO", "", "Sample ID"], "id": "KGMO Sample ID"},
            {"name": ["KGMO", "MANUAL", "FFA"], "id": "KGMO MANUAL FFA"},
            {"name": ["KGMO", "MANUAL", "CLR"], "id": "KGMO MANUAL CLR"},
            {"name": ["KGMO", "MANUAL", "MIV"], "id": "KGMO MANUAL MIV"},
            {"name": ["KGMO", "MANUAL", "EO"], "id": "KGMO MANUAL EO"},
            {"name": ["KGMO", "MANUAL", "IV"], "id": "KGMO MANUAL IV"},
            {"name": ["KGMO", "MANUAL", "SV"], "id": "KGMO MANUAL SV"},
        ]

        self.app.layout = html.Div([
            html.Link(
                rel='stylesheet',
                href='https://cdnjs.cloudflare.com/ajax/libs/bootstrap-icons/1.4.0/font/bootstrap-icons.min.css'
            ),
            html.Div([
                html.Div(className='button-container', children=[
                    html.A([html.I(className="bi bi-house"), 'Home'],
                           href='/home', className='link-button'),
                    html.Button([html.I(className="bi bi-printer"), "Print"],
                                id="print-button", className='print-button'),
                    html.Div(id='dummy-output'),
                    html.A([html.I(className="bi bi-table"), 'Daily'],
                           href='/daily_analysis_report', className='link-button'),
                    html.Button([html.I(className="bi bi-file-excel"), " Download Excel"],
                                id="p-button", className='p-button'),
                    dcc.Download(id="download-dataframe-xlsx"),

                ]),
                html.H1('Mustard Plant Process Report',
                        style={'textAlign': 'center', 'fontWeight': 'bold', 'flex': '1'}),
                html.Div([html.Label('Date:', style={
                    'margin-right': '10px', 'fontWeight': 'bold'}),
                    html.Label(datetime.now().strftime('%Y-%m-%d'),
                               style={'margin-right': '20px'}),], style={'display': 'inline', 'fontWeight': 'bold', 'border-style': 'solid',
                                                                         'border-color': 'black',
                                                                         'backgroundColor': 'white',
                                                                         'padding': '5px 10px',
                                                                         'borderRadius': '5px',
                                                                         'fontSize': '20px',
                                                                         'margin': '10px'
                                                                         })
            ], style={'display': 'flex', 'align-items': 'center', 'justify-content': 'space-between', 'width': '100%', 'margin': '10px'}),
            html.Div(id='box', children=[
                html.Label('Select Report Date:', style={
                    'margin-right': '10px', 'fontWeight': 'bold'}),
                dcc.DatePickerRange(
                    id='date-picker-Range',
                    min_date_allowed=datetime(2020, 3, 28),
                    max_date_allowed=datetime.now(),
                    start_date=datetime.now().date() - timedelta(days=30),
                    end_date=datetime.now().date(),
                    display_format='YYYY-MM-DD',
                    style={
                        'display': 'inline-block',
                        'fontWeight': 'bold',
                        'backgroundColor': 'white',
                    },
                ),
            ], style={'display': 'inline-flex', 'justifyContent': 'left', 'alignItems': 'center', 'margin': '10px', 'border-style': 'solid',
                      'border-color': 'black', 'padding': '5px 10px',
                      'borderRadius': '5px', 'fontSize': '20px',
                      'margin': '10px'}),
            dash_table.DataTable(
                id='table',
                columns=columns,
                data=reshaped_data,
                merge_duplicate_headers=True,
                page_size=10000,
                style_cell={
                    'textAlign': 'center',
                    'minWidth': '10px',
                    'width': '50px',
                    'maxWidth': '150px',
                    'whiteSpace': 'normal',
                },
                style_header={
                    'backgroundColor': 'rgb(230, 230, 230)',
                    'fontWeight': 'bold'
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': 'rgb(248, 248, 248)'
                    },
                    {
                        'if': {'filter_query': '{Date/Time} = "Total/Average"'},
                        'fontWeight': 'bold',
                        'backgroundColor': 'rgb(230, 230, 230)'
                    }
                ]
            )
        ])

    def setup_callbacks(self):
        self.app.clientside_callback(
            ClientsideFunction(
                namespace='clientside',
                function_name='printPage'
            ),
            Output('dummy-output', 'children'),
            [Input('print-button', 'n_clicks')]
        )

        @self.app.callback(
            Output("download-dataframe-xlsx", "data"),
            Input("p-button", "n_clicks"),
            Input('date-picker-Range', 'start_date'),
            Input('date-picker-Range', 'end_date'),
            prevent_initial_call=True,
        )
        def func(n_clicks,start_date, end_date):
            reshaped_data = update_table(start_date, end_date)
            formatted_df = self.format_dataframe_for_excel(reshaped_data)
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Daily Report"

            for r_idx, row in enumerate(dataframe_to_rows(formatted_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

            for col in range(1, len(formatted_df.columns) + 1):
                col_letter = get_column_letter(col)
                header = formatted_df.columns[col - 1]

                worksheet[f'{col_letter}3'].value = header[2]
                worksheet[f'{col_letter}3'].font = Font(bold=True)

                if header[1]:
                    worksheet[f'{col_letter}2'].value = header[1]
                    worksheet[f'{col_letter}2'].font = Font(bold=True)

                if header[0]:
                    worksheet[f'{col_letter}1'].value = header[0]
                    worksheet[f'{col_letter}1'].font = Font(bold=True)

            primary_headers = ["Mustard Seed",
                               "Mustard Cake", "Mustard Ghani", "KGMO"]
            for header in primary_headers:
                start_col = formatted_df.columns.get_loc(
                    (header, '', 'Sample ID')) + 1
                end_col = start_col + \
                    sum((formatted_df.columns.get_level_values(0) == header)) - 1
                if start_col != end_col:
                    start_col_letter = get_column_letter(start_col)
                    end_col_letter = get_column_letter(end_col)
                    worksheet.merge_cells(
                        f'{start_col_letter}1:{end_col_letter}1')

            for col in range(1, len(formatted_df.columns) + 1):
                if formatted_df.columns[col - 1][1]:
                    start_col = col
                    while col <= len(formatted_df.columns) and formatted_df.columns[col - 1][1] == formatted_df.columns[start_col - 1][1]:
                        col += 1
                    end_col = col - 1
                    if start_col != end_col:
                        start_col_letter = get_column_letter(start_col)
                        end_col_letter = get_column_letter(end_col)
                        worksheet.merge_cells(
                            f'{start_col_letter}2:{end_col_letter}2')

            workbook.save(output)
            output.seek(0)
            return dcc.send_bytes(output.getvalue(), "daily_report.xlsx")

        @self.app.callback(
            Output('table', 'data'),
            [
                Input('date-picker-Range', 'start_date'),
                Input('date-picker-Range', 'end_date')
            ]
        )
        def update_table(start_date, end_date):
            data = self.get_data()
            reshaped_data = self.reshape_data(data)

            if start_date is not None and end_date is not None:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(
                    end_date, '%Y-%m-%d') + timedelta(days=1)

                filtered_data = [
                    row for row in reshaped_data
                    if row['Date/Time'] != 'Total/Average' and
                    start_date <= datetime.strptime(
                        row['Date/Time'], '%Y-%m-%d %H:%M') < end_date
                ]
                filtered_data = self.append_footer_data(filtered_data)
                return filtered_data

            return reshaped_data
